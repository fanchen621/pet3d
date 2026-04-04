"""
宠物大冒险 - Flask Backend Server
支持班级优化大师学生导入 + 积分深度联动
"""

import json
import random
import os
import sys
import io
import uuid
from datetime import datetime
from flask import Flask, render_template, jsonify, request, session
from werkzeug.utils import secure_filename

# Ensure working directory is this file's directory (for direct run from PyCharm)
_app_dir = os.path.dirname(os.path.abspath(__file__))
if _app_dir not in sys.path:
    sys.path.insert(0, _app_dir)
os.chdir(_app_dir)

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("⚠️  警告: openpyxl 未安装，Excel导入功能将不可用。请运行: pip install openpyxl")

import database as db
import models

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.secret_key = os.environ.get('SECRET_KEY', 'pet3d-secret-key-change-in-prod')

# Initialize database on startup
db.init_db()
db.create_player()

# Per-session battle states
BATTLE_STATES = {}

SHOP_ITEMS = {
    'food': [
        {'id': 'apple', 'name': '魔法苹果', 'price': 10, 'type': 'food', 'effect': {'mood': 20}},
        {'id': 'cake', 'name': '星光蛋糕', 'price': 30, 'type': 'food', 'effect': {'mood': 50}},
        {'id': 'feast', 'name': '龙之盛宴', 'price': 80, 'type': 'food', 'effect': {'mood': 100}},
        {'id': 'candy', 'name': '彩虹糖果', 'price': 5, 'type': 'food', 'effect': {'mood': 10}},
        {'id': 'potion', 'name': '经验药水', 'price': 50, 'type': 'food', 'effect': {'exp': 50}},
        {'id': 'elixir', 'name': '超级经验药水', 'price': 150, 'type': 'food', 'effect': {'exp': 200}},
    ],
    'costumes': [
        {'id': 'hat_crown', 'name': '小皇冠', 'price': 100, 'type': 'costume', 'slot': 'hat', 'rarity': 'rare'},
        {'id': 'hat_wizard', 'name': '巫师帽', 'price': 120, 'type': 'costume', 'slot': 'hat', 'rarity': 'rare'},
        {'id': 'hat_flower', 'name': '花环', 'price': 60, 'type': 'costume', 'slot': 'hat', 'rarity': 'uncommon'},
        {'id': 'acc_glasses', 'name': '酷炫墨镜', 'price': 80, 'type': 'costume', 'slot': 'accessory', 'rarity': 'uncommon'},
        {'id': 'acc_wings', 'name': '精灵翅膀', 'price': 200, 'type': 'costume', 'slot': 'accessory', 'rarity': 'epic'},
        {'id': 'acc_cape', 'name': '英雄披风', 'price': 150, 'type': 'costume', 'slot': 'accessory', 'rarity': 'epic'},
    ],
    'skills': [
        {'id': 'scroll_fire', 'name': '火焰技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'fire', 'rarity': 'rare'},
        {'id': 'scroll_ice', 'name': '冰霜技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'ice', 'rarity': 'rare'},
        {'id': 'scroll_elec', 'name': '雷电技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'electric', 'rarity': 'rare'},
        {'id': 'scroll_nature', 'name': '自然技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'nature', 'rarity': 'rare'},
        {'id': 'scroll_dark', 'name': '暗影技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'dark', 'rarity': 'rare'},
        {'id': 'scroll_light', 'name': '光明技能卷轴', 'price': 100, 'type': 'skill_scroll', 'element': 'light', 'rarity': 'rare'},
    ],
    'evolution': [
        {'id': 'stone_fire', 'name': '火焰进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'fire', 'rarity': 'epic'},
        {'id': 'stone_ice', 'name': '冰霜进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'ice', 'rarity': 'epic'},
        {'id': 'stone_elec', 'name': '雷电进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'electric', 'rarity': 'epic'},
        {'id': 'stone_nature', 'name': '自然进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'nature', 'rarity': 'epic'},
        {'id': 'stone_dark', 'name': '暗影进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'dark', 'rarity': 'epic'},
        {'id': 'stone_light', 'name': '光明进化石', 'price': 300, 'type': 'evolution_stone', 'element': 'light', 'rarity': 'epic'},
        {'id': 'super_stone', 'name': '超级进化石', 'price': 800, 'type': 'evolution_stone', 'element': 'any', 'rarity': 'legendary'},
    ],
    'special': [
        {'id': 'rename', 'name': '改名卡', 'price': 50, 'type': 'special', 'rarity': 'common'},
        {'id': 'revive', 'name': '复活药水', 'price': 40, 'type': 'special', 'rarity': 'uncommon'},
        {'id': 'exp_boost', 'name': '经验加倍卡', 'price': 60, 'type': 'special', 'rarity': 'uncommon'},
        {'id': 'lucky', 'name': '幸运符', 'price': 70, 'type': 'special', 'rarity': 'uncommon'},
        {'id': 'capture', 'name': '宠物球', 'price': 100, 'type': 'special', 'rarity': 'rare'},
    ]
}

# Battle items that can be used during combat
BATTLE_ITEMS = {
    'potion_battle': {'name': '战斗药水', 'icon': '🧪', 'effect': {'heal': 0.3}, 'desc': '恢复30%生命'},
    'potion_super': {'name': '超级药水', 'icon': '💊', 'effect': {'heal': 0.6}, 'desc': '恢复60%生命'},
    'capture': {'name': '宠物球', 'icon': '🔮', 'effect': {'capture': True}, 'desc': '尝试捕捉宠物'},
    'revive': {'name': '复活药水', 'icon': '💖', 'effect': {'revive': 0.5}, 'desc': '复活并恢复50%HP'},
}

ACHIEVEMENTS = {
    'first_pet': {'name': '初遇伙伴', 'desc': '获得第一只宠物', 'icon': '🐾'},
    'first_win': {'name': '初战告捷', 'desc': '赢得第一场战斗', 'icon': '⚔️'},
    'first_evolution': {'name': '进化之路', 'desc': '宠物第一次进化', 'icon': '✨'},
    'super_evolution': {'name': '超级进化', 'desc': '宠物超进化', 'icon': '🌟'},
    'level_10': {'name': '小有成就', 'desc': '宠物达到10级', 'icon': '📈'},
    'level_50': {'name': '大师之路', 'desc': '宠物达到50级', 'icon': '🏆'},
    'first_purchase': {'name': '购物达人', 'desc': '第一次在商店购买物品', 'icon': '🛒'},
    'collector_5': {'name': '收藏家', 'desc': '收集5只宠物', 'icon': '📚'},
    'cp_1000': {'name': '战力千钧', 'desc': '宠物战力达到1000', 'icon': '💪'},
    'cp_5000': {'name': '战力无双', 'desc': '宠物战力达到5000', 'icon': '🔥'},
}

BONUS_SKILLS = {
    'fire': [
        {'name': '烈焰新星', 'power': 50, 'element': 'fire', 'unlock_level': 1},
        {'name': '地狱火', 'power': 70, 'element': 'fire', 'unlock_level': 1},
    ],
    'ice': [
        {'name': '冰锥术', 'power': 48, 'element': 'ice', 'unlock_level': 1},
        {'name': '寒冰护盾', 'power': 0, 'element': 'ice', 'unlock_level': 1, 'buff': 'defense'},
    ],
    'electric': [
        {'name': '电磁炮', 'power': 55, 'element': 'electric', 'unlock_level': 1},
        {'name': '雷电领域', 'power': 65, 'element': 'electric', 'unlock_level': 1},
    ],
    'nature': [
        {'name': '荆棘缠绕', 'power': 45, 'element': 'nature', 'unlock_level': 1},
        {'name': '生命汲取', 'power': 40, 'element': 'nature', 'unlock_level': 1, 'heal': True},
    ],
    'dark': [
        {'name': '虚空裂隙', 'power': 60, 'element': 'dark', 'unlock_level': 1},
        {'name': '暗影突袭', 'power': 50, 'element': 'dark', 'unlock_level': 1},
    ],
    'light': [
        {'name': '圣光审判', 'power': 55, 'element': 'light', 'unlock_level': 1},
        {'name': '天使庇护', 'power': 0, 'element': 'light', 'unlock_level': 1, 'heal': True},
    ],
}

ELEMENT_ADVANTAGE = {
    'fire': 'nature', 'nature': 'ice', 'ice': 'fire',
    'electric': 'light', 'light': 'dark', 'dark': 'electric',
}

# Elemental reactions: (element1, element2) -> reaction info
ELEMENTAL_REACTIONS = {
    ('fire', 'ice'): {'name': '蒸汽', 'bonus_damage': 15, 'desc': '水火相遇产生蒸汽！'},
    ('ice', 'fire'): {'name': '蒸汽', 'bonus_damage': 15, 'desc': '水火相遇产生蒸汽！'},
    ('fire', 'electric'): {'name': '等离子', 'bonus_damage': 20, 'desc': '电火交融产生等离子！'},
    ('electric', 'fire'): {'name': '等离子', 'bonus_damage': 20, 'desc': '电火交融产生等离子！'},
    ('ice', 'electric'): {'name': '超导', 'bonus_damage': 12, 'desc': '冰电共鸣产生超导！'},
    ('electric', 'ice'): {'name': '超导', 'bonus_damage': 12, 'desc': '冰电共鸣产生超导！'},
    ('nature', 'fire'): {'name': '燃烧', 'bonus_damage': 10, 'desc': '自然之力引发燃烧！'},
    ('dark', 'light'): {'name': '湮灭', 'bonus_damage': 25, 'desc': '光暗碰撞产生湮灭！'},
    ('light', 'dark'): {'name': '湮灭', 'bonus_damage': 25, 'desc': '光暗碰撞产生湮灭！'},
    ('nature', 'ice'): {'name': '冰冻植物', 'bonus_damage': 8, 'desc': '植物被冰冻覆盖！'},
}


def get_battle_state():
    """Get battle state for current session."""
    sid = session.get('battle_id')
    return BATTLE_STATES.get(sid) if sid else None


def set_battle_state(state):
    """Set battle state for current session."""
    sid = session.get('battle_id')
    if not sid:
        sid = str(uuid.uuid4())
        session['battle_id'] = sid
    if state is None:
        BATTLE_STATES.pop(sid, None)
    else:
        BATTLE_STATES[sid] = state
    return state


def find_shop_item(item_id):
    for cat in SHOP_ITEMS.values():
        for item in cat:
            if item['id'] == item_id:
                return item
    return None


def check_achievements():
    player = db.get_player()
    pets = db.get_all_pets()
    inventory = db.get_inventory()
    existing = {a['achievement_id'] for a in db.get_achievements()}
    new_achievements = []

    def unlock(aid):
        if aid not in existing and db.unlock_achievement(1, aid):
            new_achievements.append(ACHIEVEMENTS[aid])

    if pets:
        unlock('first_pet')
    if len(pets) >= 5:
        unlock('collector_5')
    if inventory:
        unlock('first_purchase')

    active = db.get_active_pet()
    if active:
        if active['level'] >= 10:
            unlock('level_10')
        if active['level'] >= 50:
            unlock('level_50')
        if active['evolution'] >= 1:
            unlock('first_evolution')
        if active['evolution'] >= 2:
            unlock('super_evolution')
        if active.get('combat_power', 0) >= 1000:
            unlock('cp_1000')
        if active.get('combat_power', 0) >= 5000:
            unlock('cp_5000')

    history = db.get_battle_history()
    if any(h['won'] for h in history):
        unlock('first_win')

    return new_achievements


# ---- Routes ----

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/state')
def api_state():
    player = db.get_player()
    pet = db.get_active_pet()
    pets = db.get_all_pets()
    inventory = db.get_inventory()
    achievements = db.get_achievements()
    students = db.get_all_students()
    classroom_stats = db.get_classroom_stats()
    return jsonify({
        'player': player,
        'pet': pet,
        'pets': pets,
        'inventory': inventory,
        'achievements': achievements,
        'students': students,
        'classroom_stats': classroom_stats,
    })


@app.route('/api/update_player', methods=['POST'])
def api_update_player():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '名字不能为空'})

    player = db.update_player(name=name)
    return jsonify({'success': True, 'player': player})


@app.route('/api/add_points', methods=['POST'])
def api_add_points():
    data = request.json
    points = int(data.get('points', 0))
    if points <= 0:
        return jsonify({'success': False, 'message': '积分必须大于0'})

    player = db.update_player(points_delta=points)

    # Check achievements
    new_ach = check_achievements()

    return jsonify({
        'success': True,
        'total_points': player['points'],
        'new_achievements': new_ach,
    })


# ===== EXCEL IMPORT (班级优化大师) =====

@app.route('/api/import_excel', methods=['POST'])
def api_import_excel():
    """Import students from 班级优化大师 Excel export."""
    if openpyxl is None:
        return jsonify({'success': False, 'message': '服务器未安装openpyxl，请联系管理员运行: pip install openpyxl'})

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})

    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx 和 .csv 文件（.xls 格式请先用Excel另存为 .xlsx）'})

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active

        students = []
        # Try to detect header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
            row_str = ' '.join(str(c or '').lower() for c in row)
            if any(k in row_str for k in ['姓名', 'name', '学生', '名字', '编号', '序号', '学号']):
                header_row = row_idx
                headers = [str(c or '').strip().lower() for c in row]
                break

        if header_row is None:
            # Assume row 1 is header
            header_row = 1
            headers = [str(c or '').strip().lower() for c in next(ws.iter_rows(max_row=1, values_only=True))]

        # Find column indices
        name_col = None
        no_col = None
        gender_col = None
        points_col = None

        for i, h in enumerate(headers):
            h_clean = h.strip()
            if any(k in h_clean for k in ['姓名', '名字', '学生姓名', '学生名', 'name', '学生']):
                name_col = i
            elif any(k in h_clean for k in ['学号', '编号', '序号', 'no', 'number', '座号', '座位']):
                no_col = i
            elif any(k in h_clean for k in ['性别', 'gender', 'sex']):
                gender_col = i
            elif any(k in h_clean for k in ['积分', '分数', 'point', 'score', '评语积分', '总分']):
                points_col = i

        # Default: if no name column found, try the first column that has text values
        if name_col is None:
            name_col = 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row is None or all(c is None for c in row):
                continue

            name = str(row[name_col] or '').strip() if name_col is not None and name_col < len(row) else ''
            if not name or name in ('', 'None'):
                continue

            student_no = str(row[no_col] or '').strip() if no_col is not None and no_col < len(row) else ''
            gender = str(row[gender_col] or '').strip() if gender_col is not None and gender_col < len(row) else ''
            points = 0
            if points_col is not None and points_col < len(row) and row[points_col] is not None:
                try:
                    points = int(float(row[points_col]))
                except (ValueError, TypeError):
                    points = 0

            students.append({
                'name': name,
                'student_no': student_no,
                'gender': gender,
                'points': 0,  # 强制初始积分为0
            })

        if not students:
            return jsonify({'success': False, 'message': '未找到学生数据，请检查文件格式'})

        result = db.import_students(students)

        return jsonify({
            'success': True,
            'imported': result['imported'],
            'updated': result['updated'],
            'total': len(students),
            'students': db.get_all_students(),
            'classroom_stats': db.get_classroom_stats(),
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})


# ===== STUDENT MANAGEMENT =====

@app.route('/api/students')
def api_get_students():
    students = db.get_all_students()
    stats = db.get_classroom_stats()
    return jsonify({'students': students, 'classroom_stats': stats})


@app.route('/api/student/<int:student_id>')
def api_get_student(student_id):
    student = db.get_student(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'})
    logs = db.get_point_logs(student_id)
    return jsonify({'success': True, 'student': student, 'point_logs': logs})


@app.route('/api/student/<int:student_id>/points', methods=['POST'])
def api_update_student_points(student_id):
    """Add or subtract points for a student. Syncs to pet mood."""
    data = request.json
    delta = int(data.get('delta', 0))
    reason = data.get('reason', '')

    if delta == 0:
        return jsonify({'success': False, 'message': '积分变化不能为0'})

    student = db.get_student(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'})

    if student['current_points'] + delta < 0:
        return jsonify({'success': False, 'message': '积分不足，无法扣除'})

    student = db.update_student_points(student_id, delta, reason)

    return jsonify({
        'success': True,
        'student': student,
        'classroom_stats': db.get_classroom_stats(),
    })


@app.route('/api/student/<int:student_id>/assign_pet', methods=['POST'])
def api_assign_pet(student_id):
    """Assign a pet to a student."""
    data = request.json
    pet_type = data.get('pet_type')

    if pet_type not in models.PET_TYPES:
        return jsonify({'success': False, 'message': '无效的宠物类型'})

    student = db.get_student(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'})

    pet_data = db.assign_pet_to_student(student_id, pet_type)
    if pet_data:
        return jsonify({
            'success': True,
            'pet': pet_data,
            'student': db.get_student(student_id),
            'classroom_stats': db.get_classroom_stats(),
        })
    return jsonify({'success': False, 'message': '分配宠物失败'})


# ===== GACHA DRAW =====
GACHA_COST_SINGLE = 20
GACHA_COST_TEN = 180
GACHA_PET_RATE = 0.05  # 5% per draw

GACHA_ITEMS = [
    {'id': 'apple', 'name': '魔法苹果', 'emoji': '🍎', 'type': 'food', 'weight': 25},
    {'id': 'candy', 'name': '彩虹糖果', 'emoji': '🍬', 'type': 'food', 'weight': 25},
    {'id': 'potion', 'name': '经验药水', 'emoji': '🧪', 'type': 'food', 'weight': 15},
    {'id': 'hat_flower', 'name': '花环', 'emoji': '💐', 'type': 'costume', 'weight': 10},
    {'id': 'acc_glasses', 'name': '酷炫墨镜', 'emoji': '🕶️', 'type': 'costume', 'weight': 8},
    {'id': 'scroll_fire', 'name': '火焰技能卷轴', 'emoji': '📜', 'type': 'skill', 'weight': 5},
    {'id': 'scroll_ice', 'name': '冰霜技能卷轴', 'emoji': '📜', 'type': 'skill', 'weight': 5},
    {'id': 'scroll_elec', 'name': '雷电技能卷轴', 'emoji': '📜', 'type': 'skill', 'weight': 5},
    {'id': 'scroll_nature', 'name': '自然技能卷轴', 'emoji': '📜', 'type': 'skill', 'weight': 5},
    {'id': 'revive', 'name': '复活药水', 'emoji': '💖', 'type': 'special', 'weight': 7},
    {'id': 'exp_boost', 'name': '经验加倍卡', 'emoji': '⚡', 'type': 'special', 'weight': 5},
    {'id': 'lucky', 'name': '幸运符', 'emoji': '🍀', 'type': 'special', 'weight': 5},
]

PET_TYPES_LIST = list(models.PET_TYPES.keys())


def _gacha_draw_one():
    """Do one gacha draw. Returns item or pet."""
    if random.random() < GACHA_PET_RATE:
        pet_type = random.choice(PET_TYPES_LIST)
        template = models.PET_TYPES[pet_type]
        return {'type': 'pet', 'pet_type': pet_type, 'name': template['name'],
                'emoji': _PetHelper.emoji(pet_type), 'rarity': template['rarity']}
    else:
        total_w = sum(i['weight'] for i in GACHA_ITEMS)
        roll = random.randint(1, total_w)
        acc = 0
        for item in GACHA_ITEMS:
            acc += item['weight']
            if roll <= acc:
                return {'type': 'item', 'item_id': item['id'], 'name': item['name'],
                        'emoji': item['emoji'], 'item_type': item['type']}


# Helper for pet emoji (used in gacha)
class _PetHelper:
    EMOJIS = {
        'dragon': '🐉', 'fox': '🦊', 'bear': '🐻', 'rabbit': '🐰', 'cat': '🐱', 'angel': '👼',
        'phoenix': '🔥', 'krystal': '🐬', 'tiger': '🐯', 'sprite': '🌿', 'wolf': '🐺', 'unicorn': '🦄'
    }
    @staticmethod
    def emoji(t): return _PetHelper.EMOJIS.get(t, '🐾')


@app.route('/api/student/<int:student_id>/draw', methods=['POST'])
def api_draw_pet(student_id):
    """Student draws pets/items with points. Supports single (20pts) and ten-draw (180pts, pity pet)."""
    data = request.json or {}
    count = int(data.get('count', 1))
    if count not in (1, 10):
        return jsonify({'success': False, 'message': '仅支持单抽(1)或十连(10)'})

    cost = GACHA_COST_TEN if count == 10 else GACHA_COST_SINGLE

    student = db.get_student(student_id)
    if not student:
        return jsonify({'success': False, 'message': '学生不存在'})

    if student['current_points'] < cost:
        return jsonify({'success': False, 'message': f'积分不足！需要{cost}积分，当前{student["current_points"]}积分'})

    # Deduct points
    db.update_student_points(student_id, -cost, f'{"十连" if count == 10 else "单"}抽', source='gacha')

    results = []
    has_pet = False
    for i in range(count):
        roll = _gacha_draw_one()
        results.append(roll)
        if roll['type'] == 'pet':
            has_pet = True

    # Pity: ten-draw guarantees at least one pet
    if count == 10 and not has_pet:
        pet_type = random.choice(PET_TYPES_LIST)
        template = models.PET_TYPES[pet_type]
        results[9] = {'type': 'pet', 'pet_type': pet_type, 'name': template['name'],
                       'emoji': _PetHelper.emoji(pet_type), 'rarity': template['rarity']}
        has_pet = True

    # Apply results: save pets, add items to inventory
    saved_pets = []
    for r in results:
        if r['type'] == 'pet':
            pet_data = models.create_pet(r['pet_type'])
            pet_data['student_id'] = student_id
            pet_data = db.save_pet(pet_data)
            # Update student pet info
            db.assign_pet_to_student(student_id, r['pet_type'], pet_data.get('name'))
            saved_pets.append(pet_data)
        elif r['type'] == 'item':
            db.add_inventory_item(1, r['item_id'], {'name': r['name']})

    return jsonify({
        'success': True,
        'results': results,
        'saved_pets': saved_pets,
        'student': db.get_student(student_id),
        'classroom_stats': db.get_classroom_stats(),
    })


@app.route('/api/student/<int:student_id>/delete', methods=['POST'])
def api_delete_student(student_id):
    db.delete_student(student_id)
    return jsonify({
        'success': True,
        'students': db.get_all_students(),
        'classroom_stats': db.get_classroom_stats(),
    })


@app.route('/api/batch_points', methods=['POST'])
def api_batch_points():
    """Batch add/subtract points for multiple students."""
    data = request.json
    changes = data.get('changes', [])  # [{student_id, delta, reason}]
    reason = data.get('reason', '')

    results = []
    for ch in changes:
        sid = ch.get('student_id')
        delta = int(ch.get('delta', 0))
        if sid and delta != 0:
            student = db.get_student(sid)
            if student and student['current_points'] + delta >= 0:
                db.update_student_points(sid, delta, reason)
                results.append({'student_id': sid, 'success': True})

    return jsonify({
        'success': True,
        'results': results,
        'students': db.get_all_students(),
        'classroom_stats': db.get_classroom_stats(),
    })


@app.route('/api/classroom_stats')
def api_classroom_stats():
    return jsonify({'stats': db.get_classroom_stats()})


# ---- Pet Routes ----

@app.route('/api/start_pet', methods=['POST'])
def api_start_pet():
    """Choose first pet."""
    data = request.json
    pet_type = data.get('pet_type', 'dragon')

    if pet_type not in models.PET_TYPES:
        return jsonify({'success': False, 'message': '未知宠物类型'})

    existing = db.get_all_pets()
    if existing:
        return jsonify({'success': False, 'message': '你已经有宠物了'})

    pet_data = models.create_pet(pet_type)
    pet_data = db.save_pet(pet_data)

    new_ach = check_achievements()

    return jsonify({'success': True, 'pet': pet_data, 'new_achievements': new_ach})


@app.route('/api/switch_pet', methods=['POST'])
def api_switch_pet():
    data = request.json
    pet_id = data.get('pet_id')
    pet = db.get_pet_by_id(pet_id)
    if not pet:
        return jsonify({'success': False, 'message': '宠物不存在'})

    db.set_active_pet(pet_id)
    return jsonify({'success': True, 'pet': pet})


@app.route('/api/feed', methods=['POST'])
def api_feed():
    pet = db.get_active_pet()
    if not pet:
        return jsonify({'success': False, 'message': '你还没有宠物'})

    pet['mood'] = min(100, pet.get('mood', 50) + 10)
    pet = db.save_pet(pet)
    return jsonify({'success': True, 'pet': pet})


@app.route('/api/play', methods=['POST'])
def api_play():
    pet = db.get_active_pet()
    if not pet:
        return jsonify({'success': False, 'message': '你还没有宠物'})

    pet['mood'] = min(100, pet.get('mood', 50) + 15)
    result = models.gain_exp(pet, 5)
    pet = db.save_pet(result['pet'])

    new_ach = check_achievements()

    return jsonify({
        'success': True,
        'pet': pet,
        'leveled_up': result['leveled_up'],
        'new_achievements': new_ach,
    })


@app.route('/api/add_mood', methods=['POST'])
def api_add_mood():
    data = request.json
    amount = int(data.get('amount', 0))
    pet = db.get_active_pet()
    if pet:
        pet['mood'] = max(0, min(100, pet.get('mood', 50) + amount))
        db.save_pet(pet)
    return jsonify({'success': True})


# ---- Battle ----

@app.route('/api/start_battle', methods=['POST'])
def api_start_battle():
    data = request.json or {}
    mode = data.get('mode', 'pve')

    pet = db.get_active_pet()
    if not pet:
        # Try to get any pet or create one
        all_pets = db.get_all_pets()
        if all_pets:
            pet = all_pets[0]
            db.set_active_pet(pet['id'])
        else:
            return jsonify({'success': False, 'message': '你还没有宠物，先去选择一只吧！'})

    # Ensure pet has required fields
    if pet.get('type') not in models.PET_TYPES:
        return jsonify({'success': False, 'message': '宠物数据异常，请重新选择宠物'})

    if pet.get('hp', 0) <= 0:
        pet['hp'] = pet.get('max_hp', 100)
        db.save_pet(pet)

    player_level = pet.get('level', 1)
    enemy = models.create_wild_pet(player_level)

    # Ensure enemy has required fields
    if not enemy.get('type') or enemy['type'] not in models.PET_TYPES:
        enemy['type'] = 'dragon'
        enemy['name'] = '野生龙宝宝'

    # Ensure enemy has skills
    if not enemy.get('skills'):
        template = models.PET_TYPES.get(enemy['type'], {})
        enemy['skills'] = [s for s in template.get('skills', []) if s.get('unlock_level', 1) <= enemy.get('level', 1)]

    # Ensure enemy stats are valid
    for stat in ('hp', 'max_hp', 'attack', 'defense', 'special', 'speed'):
        if not enemy.get(stat) or enemy[stat] <= 0:
            template = models.PET_TYPES.get(enemy['type'], {})
            defaults = {'hp': 100, 'max_hp': 100, 'attack': 20, 'defense': 15, 'special': 20, 'speed': 15}
            base_key = f'base_{stat}'
            enemy[stat] = template.get(base_key, defaults.get(stat, 10))
            if stat == 'hp':
                enemy['max_hp'] = enemy[stat]

    state = {
        'player_pet': pet,
        'enemy_pet': enemy,
        'mode': mode,
        'turn': 0,
        'combo_count': 0,
        'last_element': None,
        'skill_cooldowns': {},
        'player_defending': False,
    }
    set_battle_state(state)

    return jsonify({
        'success': True,
        'player_pet': pet,
        'enemy_pet': enemy,
        'mode': mode,
    })


@app.route('/api/battle_action', methods=['POST'])
def api_battle_action():
    data = request.json
    action = data.get('action', 'attack')
    skill_index = data.get('skill_index', 0)

    state = get_battle_state()
    if not state:
        # Battle state lost (session expired) - allow starting fresh
        return jsonify({
            'success': False,
            'message': '战斗数据已过期，请重新开始战斗',
            'error_code': 'battle_state_lost'
        })

    player_pet = state['player_pet']
    enemy_pet = state['enemy_pet']
    state['turn'] += 1

    logs = []
    player_damage = 0
    enemy_damage = 0
    battle_over = False
    player_won = False
    combo_count = state.get('combo_count', 0)
    reactions = []

    def get_multiplier(attacker_elem, defender_elem):
        if ELEMENT_ADVANTAGE.get(attacker_elem) == defender_elem:
            return 1.5
        if ELEMENT_ADVANTAGE.get(defender_elem) == attacker_elem:
            return 0.75
        return 1.0

    def check_elemental_reaction(attack_elem, target_elem):
        """Check if two elements create a reaction."""
        reaction = ELEMENTAL_REACTIONS.get((attack_elem, target_elem))
        return reaction

    def apply_damage(target_pet, base_damage, attack_elem, attacker_name, is_special=False):
        nonlocal player_damage, enemy_damage, combo_count, reactions

        target_elem = models.PET_TYPES[target_pet['type']]['element']

        # Check elemental reaction
        reaction = check_elemental_reaction(attack_elem, target_elem)
        bonus_dmg = 0
        if reaction:
            bonus_dmg = reaction['bonus_damage']
            reactions.append(reaction)

        # Combo bonus
        combo_mult = 1.0
        if combo_count > 0:
            combo_mult = 1.0 + combo_count * 0.1
            combo_mult = min(combo_mult, 2.0)  # Cap at 2x

        crit = random.random() < 0.12
        dmg = max(1, int(base_damage * combo_mult * (0.85 + random.random() * 0.3)))
        if crit:
            dmg = int(dmg * 1.8)
        dmg += bonus_dmg

        target_pet['hp'] = max(0, target_pet['hp'] - dmg)

        if target_pet == enemy_pet:
            player_damage = dmg
        else:
            enemy_damage = dmg

        combo_count += 1
        state['last_element'] = attack_elem

        log = {
            'text': f"{attacker_name} 造成 {dmg} 点伤害{'（暴击！）' if crit else ''}{' ' + reaction['desc'] if reaction else ''}",
            'type': 'player' if target_pet == enemy_pet else 'enemy',
            'damage': dmg, 'crit': crit,
            'element': attack_elem,
            'combo': combo_count if combo_count > 1 else 0,
            'reaction': reaction['name'] if reaction else None,
        }
        return log

    if action == 'attack':
        base = player_pet['attack']
        mult = get_multiplier(
            models.PET_TYPES[player_pet['type']]['element'],
            models.PET_TYPES[enemy_pet['type']]['element']
        )
        elem = models.PET_TYPES[player_pet['type']]['element']
        log = apply_damage(enemy_pet, (base - enemy_pet['defense'] * 0.5) * mult, elem, player_pet['name'])
        logs.append(log)

    elif action == 'special':
        skills = player_pet.get('skills', [])
        if skill_index < len(skills):
            skill = skills[skill_index]
            skill_name = skill['name']

            # Check cooldown
            cooldowns = state.get('skill_cooldowns', {})
            cd = cooldowns.get(skill_name, 0)
            if cd > 0:
                logs.append({'text': f"{skill_name} 冷却中（还需{cd}回合）", 'type': 'info'})
            else:
                if skill.get('heal'):
                    heal = int(player_pet['max_hp'] * 0.3)
                    player_pet['hp'] = min(player_pet['max_hp'], player_pet['hp'] + heal)
                    combo_count = 0  # Healing resets combo
                    logs.append({'text': f"{player_pet['name']} 使用了 {skill_name}！恢复 {heal} HP", 'type': 'heal', 'heal': heal})
                else:
                    base = skill['power'] + player_pet['special'] * 0.5
                    mult = get_multiplier(skill['element'], models.PET_TYPES[enemy_pet['type']]['element'])
                    log = apply_damage(enemy_pet, (base - enemy_pet['defense'] * 0.3) * mult, skill['element'], player_pet['name'])
                    logs.append(log)

                # Set cooldown (stronger skills = longer cooldown)
                cd_turns = 1 if skill['power'] < 50 else 2 if skill['power'] < 70 else 3
                cooldowns[skill_name] = cd_turns
                state['skill_cooldowns'] = cooldowns
        else:
            logs.append({'text': '没有可使用的技能！', 'type': 'info'})

    elif action == 'defend':
        logs.append({'text': f"{player_pet['name']} 进入防御姿态！", 'type': 'defend'})
        state['player_defending'] = True
        combo_count = 0

    elif action == 'use_item':
        item_id = data.get('item_id', '')
        item_info = BATTLE_ITEMS.get(item_id)
        if item_info:
            effect = item_info.get('effect', {})
            if effect.get('heal'):
                heal = int(player_pet['max_hp'] * effect['heal'])
                player_pet['hp'] = min(player_pet['max_hp'], player_pet['hp'] + heal)
                logs.append({'text': f"使用了 {item_info['name']}！恢复 {heal} HP", 'type': 'heal', 'heal': heal})
            elif effect.get('revive'):
                if player_pet['hp'] <= 0:
                    player_pet['hp'] = int(player_pet['max_hp'] * effect['revive'])
                    logs.append({'text': f"使用了 {item_info['name']}！宠物复活了！", 'type': 'heal', 'heal': player_pet['hp']})
            elif effect.get('capture'):
                # Capture mechanic: success depends on enemy HP
                enemy_pct = enemy_pet['hp'] / enemy_pet['max_hp']
                capture_chance = (1 - enemy_pct) * 0.6 + 0.1  # 10% at full HP, 70% at 0 HP
                if enemy_pet['level'] > player_pet['level']:
                    capture_chance *= 0.5
                if random.random() < capture_chance:
                    battle_over = True
                    player_won = True
                    # Add wild pet to collection
                    captured_pet = enemy_pet.copy()
                    captured_pet['name'] = captured_pet['name'].replace('野生', '')
                    captured_pet['hp'] = captured_pet['max_hp']
                    captured_pet.pop('is_wild', None)
                    captured_pet.pop('id', None)
                    captured_pet = db.save_pet(captured_pet)
                    logs.append({'text': f"🎉 成功捕获了 {enemy_pet['name']}！", 'type': 'victory', 'captured': True})
                    result = {
                        'success': True, 'logs': logs,
                        'player_hp': player_pet['hp'], 'player_max_hp': player_pet['max_hp'],
                        'enemy_hp': 0, 'enemy_max_hp': enemy_pet['max_hp'],
                        'player_damage': 0, 'enemy_damage': 0,
                        'battle_over': True, 'player_won': True, 'turn': state['turn'],
                        'pet': db.get_active_pet(),
                        'captured_pet': captured_pet,
                        'reactions': reactions,
                        'combo_count': combo_count,
                    }
                    set_battle_state(None)
                    return jsonify(result)
                else:
                    logs.append({'text': f"捕获失败！{enemy_pet['name']} 挣脱了！", 'type': 'info'})
            # Consume item from inventory
            db.use_inventory_item(1, item_id)
        else:
            logs.append({'text': '没有这个道具！', 'type': 'info'})

    elif action == 'flee':
        if random.random() < 0.6:
            logs.append({'text': '成功逃跑了！', 'type': 'flee'})
            battle_over = True
        else:
            logs.append({'text': '逃跑失败！', 'type': 'info'})
            combo_count = 0

    if enemy_pet['hp'] <= 0 and not battle_over:
        battle_over = True
        player_won = True
        exp_gain = int(20 + enemy_pet['level'] * 5 + random.randint(0, 10))
        logs.append({'text': f"🎉 胜利！获得 {exp_gain} 经验值！", 'type': 'victory', 'exp': exp_gain})
    elif not battle_over:
        # Decrement cooldowns for enemy skills
        enemy_skills = enemy_pet.get('skills', [])
        if enemy_skills and random.random() < 0.4:
            skill = random.choice(enemy_skills)
            if skill.get('heal'):
                heal = int(enemy_pet['max_hp'] * 0.25)
                enemy_pet['hp'] = min(enemy_pet['max_hp'], enemy_pet['hp'] + heal)
                logs.append({'text': f"{enemy_pet['name']} 使用了 {skill['name']}！恢复 {heal} HP", 'type': 'enemy_heal', 'heal': heal})
            else:
                base = skill['power'] + enemy_pet['special'] * 0.5
                mult = get_multiplier(skill['element'], models.PET_TYPES[player_pet['type']]['element'])
                dmg = max(1, int((base - player_pet['defense'] * 0.3) * mult * (0.85 + random.random() * 0.3)))
                if state.get('player_defending'):
                    dmg = int(dmg * 0.5)
                player_pet['hp'] = max(0, player_pet['hp'] - dmg)
                enemy_damage = dmg
                # Check reaction from enemy element
                reaction = check_elemental_reaction(skill['element'], models.PET_TYPES[player_pet['type']]['element'])
                reaction_text = f" {reaction['desc']}" if reaction else ""
                logs.append({'text': f"{enemy_pet['name']} 使用了 {skill['name']}！造成 {dmg} 点伤害{reaction_text}", 'type': 'enemy', 'damage': dmg, 'element': skill['element']})
        else:
            base = enemy_pet['attack']
            mult = get_multiplier(models.PET_TYPES[enemy_pet['type']]['element'], models.PET_TYPES[player_pet['type']]['element'])
            dmg = max(1, int((base - player_pet['defense'] * 0.5) * mult * (0.85 + random.random() * 0.3)))
            if state.get('player_defending'):
                dmg = int(dmg * 0.5)
            player_pet['hp'] = max(0, player_pet['hp'] - dmg)
            enemy_damage = dmg
            logs.append({'text': f"{enemy_pet['name']} 发起攻击！造成 {dmg} 点伤害", 'type': 'enemy', 'damage': dmg, 'element': models.PET_TYPES[enemy_pet['type']]['element']})

        if player_pet['hp'] <= 0:
            battle_over = True
            player_won = False
            exp_gain = int(5 + enemy_pet['level'] * 2)
            logs.append({'text': f"💀 失败了... 获得 {exp_gain} 经验值（安慰奖）", 'type': 'defeat', 'exp': exp_gain})

    # Decrement skill cooldowns
    cooldowns = state.get('skill_cooldowns', {})
    new_cooldowns = {}
    for skill_name, cd in cooldowns.items():
        if cd > 1:
            new_cooldowns[skill_name] = cd - 1
    state['skill_cooldowns'] = new_cooldowns

    state['player_defending'] = False
    state['player_pet'] = player_pet
    state['enemy_pet'] = enemy_pet
    state['combo_count'] = combo_count

    result = {
        'success': True, 'logs': logs,
        'player_hp': player_pet['hp'], 'player_max_hp': player_pet['max_hp'],
        'enemy_hp': enemy_pet['hp'], 'enemy_max_hp': enemy_pet['max_hp'],
        'player_damage': player_damage, 'enemy_damage': enemy_damage,
        'battle_over': battle_over, 'player_won': player_won, 'turn': state['turn'],
        'combo_count': combo_count,
        'reactions': reactions,
        'skill_cooldowns': {k: v for k, v in state.get('skill_cooldowns', {}).items()},
    }

    if battle_over:
        exp_gain = logs[-1].get('exp', 0) if logs else 0
        level_up_info = models.gain_exp(player_pet, exp_gain)
        player_pet = db.save_pet(level_up_info['pet'])

        db.add_battle_record(1, player_pet['id'], enemy_pet['name'], enemy_pet['type'], player_won, exp_gain, state['turn'], state['mode'])

        result['pet'] = player_pet
        result['exp_gained'] = exp_gain
        result['leveled_up'] = level_up_info['leveled_up']
        result['new_achievements'] = check_achievements()

        set_battle_state(None)
    else:
        set_battle_state(state)

    return jsonify(result)


# ---- Shop ----

@app.route('/api/buy', methods=['POST'])
def api_buy():
    data = request.json
    item_id = data.get('item_id')
    price = data.get('price')
    item_data = data.get('item_data', {})

    player = db.get_player()
    if player['points'] < price:
        return jsonify({'success': False, 'message': '积分不足'})

    player = db.update_player(points_delta=-price)
    inventory = db.add_inventory_item(1, item_id, item_data)

    new_ach = check_achievements()

    return jsonify({
        'success': True,
        'remaining_points': player['points'],
        'inventory': inventory,
        'new_achievements': new_ach,
    })


@app.route('/api/use_item', methods=['POST'])
def api_use_item():
    data = request.json
    item_id = data.get('item_id')

    pet = db.get_active_pet()
    if not pet:
        return jsonify({'success': False, 'message': '你还没有宠物'})

    item = find_shop_item(item_id)
    if not item:
        return jsonify({'success': False, 'message': '未知物品'})

    if not db.use_inventory_item(1, item_id):
        return jsonify({'success': False, 'message': '物品数量不足'})

    message = '使用成功！'

    if item['type'] == 'food':
        effect = item.get('effect', {})
        if 'mood' in effect:
            pet['mood'] = min(100, pet.get('mood', 50) + effect['mood'])
            message = f"宠物心情恢复了 +{effect['mood']}！"
        if 'exp' in effect:
            result = models.gain_exp(pet, effect['exp'])
            pet = result['pet']
            message = f"获得 {effect['exp']} 经验值！"
            if result['leveled_up']:
                message += f" 升级到 Lv.{result['new_level']}！"

    elif item['type'] == 'skill_scroll':
        element = item.get('element')
        bonus = BONUS_SKILLS.get(element, [])
        if bonus:
            skill = random.choice(bonus)
            if not any(s['name'] == skill['name'] for s in pet.get('skills', [])):
                pet.setdefault('skills', []).append(skill.copy())
                message = f"学会了新技能：{skill['name']}！"
            else:
                message = '已经学会这个技能了！'

    elif item['type'] == 'evolution_stone':
        element = item.get('element')
        pet_element = models.PET_TYPES[pet['type']]['element']
        if element == 'any' or element == pet_element:
            if pet['level'] >= 20 and pet['evolution'] < 2:
                pet['evolution'] += 1
                evo_names = models.EVOLUTION_NAMES[pet['type']]
                pet['name'] = evo_names[min(pet['evolution'], len(evo_names) - 1)]
                pet['max_hp'] = int(pet['max_hp'] * 1.2)
                pet['hp'] = pet['max_hp']
                pet['attack'] = int(pet['attack'] * 1.15)
                pet['defense'] = int(pet['defense'] * 1.15)
                pet['special'] = int(pet['special'] * 1.15)
                pet['speed'] = int(pet['speed'] * 1.1)
                message = f"🎉 进化为 {pet['name']}！全属性大幅提升！"
            else:
                message = '暂时无法进化（需要 Lv.20+）'
                db.add_inventory_item(1, item_id, item)
        else:
            message = '进化石属性不匹配！'
            db.add_inventory_item(1, item_id, item)

    elif item['type'] == 'costume':
        slot = item.get('slot', 'accessory')
        pet.setdefault('costume', {})[slot] = item_id
        message = f"装备了 {item['name']}！"

    pet['combat_power'] = models.calc_combat_power(pet)
    pet = db.save_pet(pet)

    inventory = db.get_inventory()
    new_ach = check_achievements()

    return jsonify({
        'success': True,
        'pet': pet,
        'inventory': inventory,
        'message': message,
        'new_achievements': new_ach,
    })


# ---- Get inventory for battle ----
@app.route('/api/battle_items')
def api_battle_items():
    """Get items available for use in battle."""
    inventory = db.get_inventory()
    battle_usable = []
    for inv in inventory:
        item_id = inv['item_id']
        if item_id in BATTLE_ITEMS or item_id in ('potion', 'elixir', 'revive', 'capture'):
            bi = BATTLE_ITEMS.get(item_id, {})
            battle_usable.append({
                'item_id': item_id,
                'name': bi.get('name', item_id),
                'icon': bi.get('icon', '📦'),
                'desc': bi.get('desc', ''),
                'count': inv.get('count', 1),
            })
    return jsonify({'items': battle_usable})


# ---- Ranking ----

@app.route('/api/ranking')
def api_ranking():
    rankings = db.get_ranking()
    return jsonify({'ranking': rankings})


@app.route('/api/battle_history')
def api_battle_history():
    history = db.get_battle_history()
    return jsonify({'history': history})


# ---- Create Pet (for new players) ----

@app.route('/api/create_pet', methods=['POST'])
def api_create_pet():
    data = request.json
    pet_type = data.get('pet_type')

    existing_pets = db.get_all_pets()
    if existing_pets:
        return jsonify({'success': False, 'message': '已经有宠物了'})

    if pet_type not in models.PET_TYPES:
        return jsonify({'success': False, 'message': '无效的宠物类型'})

    pet_data = models.create_pet(pet_type)
    pet_data = db.save_pet(pet_data)

    new_ach = check_achievements()

    return jsonify({
        'success': True,
        'pet': pet_data,
        'new_achievements': new_ach,
    })


# ---- Reset ----

@app.route('/api/reset', methods=['POST'])
def api_reset():
    db.reset_game()
    BATTLE_STATES.clear()
    return jsonify({'success': True})


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5555, debug=False)
